"""Excel contact file reader/writer. Preserves all rows; writes status back."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.utils.logging import get_logger

log = get_logger(__name__)


# Canonical column names (case-insensitive match against the sheet header row)
REQUIRED_COLUMNS = [
    "Company Name",
    "Contact Name",
    "Contact Email",
]
ALL_COLUMNS = [
    "Company Name",
    "Contact Name",
    "Contact Email",
    "Contact Role",
    "Company Website",
    "Headcount",
    "Service Type",
    "LinkedIn Person URL",
    "LinkedIn Company URL",
    "Country",
    "Fit",
    "IT Services",
    "Outreach Status",
    "Outreach Date",
]


@dataclass
class ExcelContact:
    row_number: int  # 1-indexed including header
    company_name: str
    contact_name: str
    contact_email: str
    contact_role: str | None
    company_website: str | None
    headcount: int | None
    service_type: str | None
    linkedin_person_url: str | None
    linkedin_company_url: str | None
    country: str | None
    fit: str | None
    it_services: str | None
    outreach_status: str | None

    def is_eligible(self) -> bool:
        return (
            (self.fit or "").strip().upper() == "YES"
            and (self.it_services or "").strip().upper() == "YES"
            and not (self.outreach_status or "").strip()
        )

    def to_contact_dict(self) -> dict[str, Any]:
        return {
            "company_name": self.company_name,
            "contact_name": self.contact_name,
            "contact_email": self.contact_email.strip().lower(),
            "contact_role": self.contact_role,
            "company_website": self.company_website,
            "headcount": self.headcount,
            "service_type": self.service_type,
            "linkedin_person_url": self.linkedin_person_url,
            "linkedin_company_url": self.linkedin_company_url,
            "country": self.country,
            "excel_row_number": self.row_number,
        }


def _header_map(ws) -> dict[str, int]:
    """Build lowercase-column-name → column-index (1-based) map."""
    mapping: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            mapping[str(cell.value).strip().lower()] = col_idx
    return mapping


def _get(ws, row: int, header_map: dict[str, int], col: str) -> Any:
    idx = header_map.get(col.lower())
    if idx is None:
        return None
    return ws.cell(row=row, column=idx).value


def read_contacts(path: Path | str) -> list[ExcelContact]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Contacts file not found: {path}")

    wb = load_workbook(filename=str(path), data_only=True)
    ws = wb.active
    header_map = _header_map(ws)

    missing = [c for c in REQUIRED_COLUMNS if c.lower() not in header_map]
    if missing:
        raise ValueError(f"Excel is missing required columns: {missing}")

    contacts: list[ExcelContact] = []
    for row_idx in range(2, ws.max_row + 1):
        email = _get(ws, row_idx, header_map, "Contact Email")
        if not email:
            continue

        headcount_raw = _get(ws, row_idx, header_map, "Headcount")
        try:
            headcount = int(headcount_raw) if headcount_raw not in (None, "") else None
        except (TypeError, ValueError):
            headcount = None

        contacts.append(
            ExcelContact(
                row_number=row_idx,
                company_name=str(_get(ws, row_idx, header_map, "Company Name") or ""),
                contact_name=str(_get(ws, row_idx, header_map, "Contact Name") or ""),
                contact_email=str(email).strip(),
                contact_role=_str_or_none(_get(ws, row_idx, header_map, "Contact Role")),
                company_website=_str_or_none(_get(ws, row_idx, header_map, "Company Website")),
                headcount=headcount,
                service_type=_str_or_none(_get(ws, row_idx, header_map, "Service Type")),
                linkedin_person_url=_str_or_none(
                    _get(ws, row_idx, header_map, "LinkedIn Person URL")
                ),
                linkedin_company_url=_str_or_none(
                    _get(ws, row_idx, header_map, "LinkedIn Company URL")
                ),
                country=_str_or_none(_get(ws, row_idx, header_map, "Country")),
                fit=_str_or_none(_get(ws, row_idx, header_map, "Fit")),
                it_services=_str_or_none(_get(ws, row_idx, header_map, "IT Services")),
                outreach_status=_str_or_none(
                    _get(ws, row_idx, header_map, "Outreach Status")
                ),
            )
        )
    return contacts


def _str_or_none(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def update_row_status(
    path: Path | str,
    row_number: int,
    *,
    status: str,
    outreach_date: datetime | None = None,
) -> None:
    """Update Outreach Status (and optionally Outreach Date) for one row."""
    path = Path(path)
    wb = load_workbook(filename=str(path))
    ws = wb.active
    header_map = _header_map(ws)

    status_col = header_map.get("outreach status")
    date_col = header_map.get("outreach date")
    if status_col:
        ws.cell(row=row_number, column=status_col).value = status
    if outreach_date and date_col:
        ws.cell(row=row_number, column=date_col).value = outreach_date
    wb.save(str(path))
